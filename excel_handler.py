import openpyxl
import os

class ExcelHandler:
    def __init__(self, file_name):
        self.file_name = file_name

        if os.path.exists(self.file_name):
            self.load_excel_file()
        else:
            print(f"Excel file '{self.file_name}' does not exist.")

    def load_excel_file(self):
        """Loads the existing Excel file."""
        try:
            self.workbook = openpyxl.load_workbook(self.file_name)
            self.sheet = self.workbook.active
            print(f"Excel file '{self.file_name}' loaded.")
        except Exception as e:
            print(f"Error loading Excel file: {e}")

#----------------------------------Customers Operations---------------------#
    def load_customers(self):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb['Customers']
        customers = []
        # take 
        for row in ws.iter_rows(min_row=2, values_only=True):
            first_name, last_name, email, phone, city, description, address, sensors_code, sensors_type, sensors_description = row

            customers.append({
                "first_name": first_name,
                "last_name": last_name,
                "email": email,
                "phone": phone,
                "city": city,
                "description": description,
                "address": address,
                "sensors_code": sensors_code,
                "sensors_type": sensors_type,
                "sensor_description": sensors_description
            })
        wb.save(self.file_name)
        wb.close()
        return customers

    def save_customers(self, customer):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb['Customers']
        
        # Append the new customer's data at the end of the worksheet
        ws.append([
            customer['first_name'], 
            customer['last_name'], 
            customer['email'], 
            customer['phone'], 
            customer['city'], 
            customer['description'], 
            customer['address']
        ])

        wb.save(self.file_name)
        wb.close()

    def delete_customer(self, customer):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb['Customers']
        customer_row = self.return_customers_row(customer)
        ws.delete_rows(customer_row)
        print(customer_row)
        print(type(customer_row))
        wb.save(self.file_name)
        wb.close()

#-----------------------------------Sensors Operations---------------------#
    def save_sensor(self, customer, sensor_code, type, description):
        if description == '' or None:
            description = '*'
        if type == '' or None:
            type = '*'
        if sensor_code == '' or None:
            sensor_code = '*'

        wb = openpyxl.load_workbook(self.file_name)
        ws = wb['Customers']
        customer_row = self.return_customers_row(customer)
        current_sensor_codes = ws[f"H{customer_row}"].value 
        current_sensor_types = ws[f"I{customer_row}"].value 
        current_sensor_descriptions = ws[f"J{customer_row}"].value
        # Write the new Value of Sensor Code and Sensor Type and Sensor Description
        if current_sensor_codes is not None:
            ws[f"H{customer_row}"] = f"{current_sensor_codes};{sensor_code}"
        else:
            ws[f"H{customer_row}"] = sensor_code 

        if current_sensor_types is not None:
            ws[f"I{customer_row}"] = f"{current_sensor_types};{type}"
        else:
            ws[f"I{customer_row}"] = type 

        if current_sensor_descriptions is not None:
            ws[f"J{customer_row}"] = f"{current_sensor_descriptions};{description}"
        else:
            ws[f"J{customer_row}"] = description 
               
        wb.save(self.file_name)
        wb.close()

    def save_sensor_edit(self, customer, sensor_code, new_sensor_code, description):
        if new_sensor_code == '' or None:
            new_sensor_code = '*'
        else:
            new_sensor_code = str(new_sensor_code)

        if description == '' or None:
            description = '*'
        else:
            description = str(description)

        wb = openpyxl.load_workbook(self.file_name)
        ws = wb['Customers']
        customer_row = self.return_customers_row(customer)
        current_sensor_codes = ws[f"H{customer_row}"].value 
        current_sensor_descriptions = ws[f"J{customer_row}"].value

        current_sensor_codes = str(current_sensor_codes)
        current_sensor_descriptions = str(current_sensor_descriptions)


        print(f"the new sensor code : {new_sensor_code}")
        print(f"the new sensor description : {description}")
        current_sensor_descriptions = current_sensor_descriptions.split(';')
        current_sensor_codes = current_sensor_codes.split(';')

        print(current_sensor_codes)
        print(current_sensor_descriptions)

        for i, code in enumerate(current_sensor_codes):
            if code == str(sensor_code):
                current_sensor_codes[i] = new_sensor_code        # Change sensor uniqe code
                current_sensor_descriptions[i]  = description    # Change sensor description

        current_sensor_codes = ";".join(current_sensor_codes)    # Create string of sensors code with new edited code
        current_sensor_descriptions = ";".join(current_sensor_descriptions) # The same action for the descriptions

        ws[f"H{customer_row}"] = current_sensor_codes            # Rewrite all the sensors code
        ws[f"J{customer_row}"] = current_sensor_descriptions     # The same action for the descriptions!

        wb.save(self.file_name)
        wb.close()

    def load_sensors(self, customer):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb['Customers']

        # Take Customer Row from the Database
        customer_row = self.return_customers_row(customer)
        sensors_types = ws[f"I{customer_row}"].value
        sensors_code = ws[f"H{customer_row}"].value
        sensors_description = ws[f"J{customer_row}"].value

        #sensors_types = str(sensors_types)
        sensors_code = str(sensors_code)
        sensors_description = str(sensors_description)

        # Decode the sensors to the list 
        if sensors_types is not None:
            sensors_types = sensors_types.split(';')
        if sensors_code is not None:
            sensors_code = sensors_code.split(';')
        if sensors_description is not None:
            sensors_description = sensors_description.split(';')

        return sensors_types, sensors_code, sensors_description 

    def delete_sensor(self, customer, sensor_code):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb['Customers']

        # Take Customer Row from the Database
        customer_row = self.return_customers_row(customer)

        # Take Sensos string values
        sensors_types = ws[f"I{customer_row}"].value
        sensors_description = ws[f"J{customer_row}"].value
        sensors_code = ws[f"H{customer_row}"].value

        # Create a list of each sensors detail
        sensors_types = sensors_types.split(';')
        sensors_description = sensors_description.split(';')
        sensors_code = sensors_code.split(';')
        for i,code in enumerate(sensors_code):
            _t = sensors_types[i]
            _d = sensors_description[i]
            if code == sensor_code:
                print(_t, _d, code)
                sensors_code.remove(code)
                sensors_types.remove(_t)
                sensors_description.remove(_d)     

        # Create a String for each sensors values
        sensors_code = ";".join(sensors_code)
        sensors_types = ";".join(sensors_types)
        sensors_description = ";".join(sensors_description)

        # Rewrite the rest of sensors
        ws[f"H{customer_row}"] = sensors_code
        ws[f"I{customer_row}"] = sensors_types
        ws[f"J{customer_row}"] = sensors_description

        wb.save(self.file_name)
        wb.close()

    def return_customers_row(self, customer):
        from main import MyApp
        wb = openpyxl.load_workbook(MyApp.resource_path("customers_data.xlsx"))
        ws = wb["Customers"]

        last_name, first_name = customer.split(";")
        customer_row = None
        
        for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):  # Start index from 2
            if row[0] == first_name and row[1] == last_name:  # Check columns 1 and 2
                customer_row = index  # Get the row number
                #print(f"Customer {first_name} {last_name} found in row {customer_row}.")
                break
        if customer_row is None:
            print(f"Customer with the name {first_name} {last_name} is not registered in the list!")    
            return
        else:
            return customer_row

        